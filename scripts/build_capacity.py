#!/usr/bin/env python3
"""大学所在地ごとの入学者数を、設置者別（国立・公立・私立）に取り出す。

学校基本調査の同じファイルには「1 計」のほかに「2 国立」「3 私立」の表がある。
公立の表はないので、計 − 国立 − 私立 として求める。

各表の「計」列は全出身地の合計であり、都道府県に対応付けられない「その他」を含む。
つまりここで出す人数は、OD グラフの47県に閉じた数より大きい。県の受け入れ規模を
見るための数字なので、こちらを採る。
"""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

import openpyxl

from build_od import (
    PREFECTURES,
    TOTAL_LABEL,
    find_labeled_column,
    integer,
    locate_header_row,
    locate_label_column,
    normalize,
    year_from_filename,
)


# 表題の先頭にある通し番号で表を見分ける。年度により全角・半角の空白が混ざる。
TABLE_PATTERNS = {
    "total": re.compile(r"^1計"),
    "national": re.compile(r"^2国立"),
    "private": re.compile(r"^3私立"),
}


def first_sheet(workbook: openpyxl.Workbook, pattern: re.Pattern[str]) -> str:
    """該当する表の先頭シート名を返す。横分割された続きのシートは使わない。"""
    for name in workbook.sheetnames:
        sheet = workbook[name]
        for row in range(1, 7):
            for column in range(1, 5):
                label = normalize(sheet.cell(row, column).value)
                if pattern.match(label):
                    # 「(つづき)」は右半分の列だけを持ち、「計」列がない。
                    if "つづき" in label:
                        continue
                    return name
    raise ValueError(f"表が見つからない: {pattern.pattern} / {workbook.sheetnames}")


def read_totals(path: Path, pattern: re.Pattern[str]) -> dict[str, int]:
    """大学所在地ごとの入学者数（表の「計」列）を返す。"""
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    sheet = workbook[first_sheet(workbook, pattern)]
    header_row, _ = locate_header_row(sheet)
    _, destinations = locate_label_column(sheet, header_row)
    total_column = find_labeled_column(sheet, header_row, TOTAL_LABEL)
    if total_column is None:
        raise ValueError(f"「計」列が見つからない: {path.name} / {sheet.title}")

    totals = {
        destination: integer(
            sheet.cell(row, total_column).value,
            context=f"{path.name}/{sheet.title} {destination} 計",
        )
        for row, destination in destinations.items()
    }
    workbook.close()
    return totals


FIELDS = ["prefecture", "year", "entrants_total", "national", "public", "private"]


def build_year(path: Path, year: int) -> list[dict[str, object]]:
    total = read_totals(path, TABLE_PATTERNS["total"])
    national = read_totals(path, TABLE_PATTERNS["national"])
    private = read_totals(path, TABLE_PATTERNS["private"])

    rows = []
    for prefecture in PREFECTURES:
        public = total[prefecture] - national[prefecture] - private[prefecture]
        if public < 0:
            raise ValueError(
                f"公立が負になる: {prefecture} 計={total[prefecture]} "
                f"国立={national[prefecture]} 私立={private[prefecture]} ({path.name})"
            )
        rows.append({
            "prefecture": prefecture,
            "year": year,
            "entrants_total": total[prefecture],
            "national": national[prefecture],
            "public": public,
            "private": private[prefecture],
        })
    return rows


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, nargs="*", default=None)
    parser.add_argument("--output-dir", type=Path, default=Path("data/processed"))
    args = parser.parse_args()

    inputs = args.input or sorted(Path("data/raw").glob("mext-school-basic-*.xlsx"))
    if not inputs:
        raise SystemExit("入力ファイルがない")
    args.output_dir.mkdir(parents=True, exist_ok=True)

    combined: list[dict[str, object]] = []
    for path in inputs:
        year = year_from_filename(path)
        rows = build_year(path, year)
        write_csv(args.output_dir / f"capacity_{year}.csv", rows)
        combined.extend(rows)
        total = sum(int(row["entrants_total"]) for row in rows)
        national = sum(int(row["national"]) for row in rows)
        public = sum(int(row["public"]) for row in rows)
        private = sum(int(row["private"]) for row in rows)
        print(f"{year}: 入学者 {total:,} = 国立 {national:,} / 公立 {public:,} / 私立 {private:,}")

    if len(inputs) > 1:
        write_csv(args.output_dir / "capacity_all.csv", combined)
        print(f"合算: {args.output_dir}/capacity_all.csv")


if __name__ == "__main__":
    main()
