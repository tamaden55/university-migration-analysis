#!/usr/bin/env python3
"""都道府県別の高校数と大学数を、設置者別に取り出す。

学校基本調査 令和7年度の2表を使う。

- 大学: 「7 都道府県別 学校数及び学生数」
- 高校: 「122 都道府県別学校数（3-1）1．計（本校＋分校）」全日制・定時制

どちらも都道府県は行、設置者は列に並ぶ。列の位置は決め打ちせず、見出しから探す。
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl

from build_od import PREFECTURES, canonical_prefecture, integer, normalize

FOUNDERS = ("計", "国立", "公立", "私立")


def prefecture_rows(sheet, column: int) -> dict[int, str]:
    """指定列を上から見て、都道府県名の行を拾う。上部の全国計の行は名前が違うので落ちる。"""
    rows = {
        row: prefecture
        for row in range(1, sheet.max_row + 1)
        if (prefecture := canonical_prefecture(sheet.cell(row, column).value))
    }
    if len(rows) != len(PREFECTURES):
        raise ValueError(f"都道府県が47件そろわない: {sheet.title} ({len(rows)}件)")
    return rows


def founder_columns(sheet, label_row: int, columns: range, *, detail_row: int | None) -> dict[str, int]:
    """設置者ごとに、その区分の先頭列（＝その区分の計）を返す。"""
    found: dict[str, int] = {}
    for column in columns:
        label = normalize(sheet.cell(label_row, column).value)
        if label in FOUNDERS and label not in found:
            if detail_row is not None:
                # 細分がある表では、区分の先頭列が「計」になっているものだけを採る。
                detail = normalize(sheet.cell(detail_row, column).value)
                if detail not in ("計", "全日制") :
                    continue
            found[label] = column
    missing = [f for f in FOUNDERS if f not in found]
    if missing:
        raise ValueError(f"設置者の列が見つからない: {missing} ({sheet.title})")
    return found


def read_counts(path: Path, sheet_name: str, label_row: int, detail_row: int | None,
                columns: range) -> dict[str, dict[str, int]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    rows = prefecture_rows(sheet, 2)
    founders = founder_columns(sheet, label_row, columns, detail_row=detail_row)

    counts: dict[str, dict[str, int]] = {}
    for row, prefecture in rows.items():
        values = {
            founder: integer(
                sheet.cell(row, column).value,
                context=f"{path.name}/{sheet_name} {prefecture} {founder}",
            )
            for founder, column in founders.items()
        }
        breakdown = values["国立"] + values["公立"] + values["私立"]
        if breakdown != values["計"]:
            raise ValueError(
                f"設置者の内訳が計と合わない: {prefecture} "
                f"計={values['計']} 内訳={breakdown} ({path.name})"
            )
        counts[prefecture] = values
    workbook.close()
    return counts


FIELDS = [
    "prefecture", "year",
    "high_schools", "high_national", "high_public", "high_private",
    "universities", "university_national", "university_public", "university_private",
]


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--year", type=int, default=2025)
    parser.add_argument("--university", type=Path,
                        default=Path("data/raw/mext-university-count-2025.xlsx"))
    parser.add_argument("--high-school", type=Path,
                        default=Path("data/raw/mext-highschool-count-2025.xlsx"))
    parser.add_argument("--output", type=Path, default=Path("data/processed/schools_2025.csv"))
    args = parser.parse_args()

    # 大学表は 4行目が「学校数 / 学生数」、6行目が設置者。学校数の区分だけを見る。
    universities = read_counts(args.university, "07", label_row=6, detail_row=None,
                               columns=range(3, 7))
    # 高校表は 4行目が設置者、6行目が課程。区分の先頭列を採る。
    high_schools = read_counts(args.high_school, "122(3-1)", label_row=4, detail_row=6,
                               columns=range(3, 40))

    rows = []
    for prefecture in PREFECTURES:
        high = high_schools[prefecture]
        university = universities[prefecture]
        rows.append({
            "prefecture": prefecture,
            "year": args.year,
            "high_schools": high["計"],
            "high_national": high["国立"],
            "high_public": high["公立"],
            "high_private": high["私立"],
            "universities": university["計"],
            "university_national": university["国立"],
            "university_public": university["公立"],
            "university_private": university["私立"],
        })

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    print(f"{args.output}: 高校 {sum(r['high_schools'] for r in rows):,}校 "
          f"/ 大学 {sum(r['universities'] for r in rows):,}校")


if __name__ == "__main__":
    main()
