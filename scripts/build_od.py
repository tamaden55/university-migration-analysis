#!/usr/bin/env python3
"""学校基本調査の都道府県間 OD 表を、グラフ DB 用のノード・エッジ CSV に変換する。

年度によって表のレイアウトが違う。

- 2020年度以前: 「1 計」表が2シートに横分割される（北海道〜愛知 / 三重〜沖縄）。
  県名は「青　森」のように全角スペース入りで、都府県の接尾辞が付かない。
- 2021年度以降: 「1 計」表が1シートに収まり、県名は「青森県」形式。
  ただし2021年度だけ行見出しが接尾辞なしで、列見出しだけ接尾辞付き。

そのためセル位置は決め打ちせず、県名の一致でヘッダ行と見出し列を検出する。
"""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl


PREFECTURES = (
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県", "茨城県",
    "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県", "新潟県", "富山県",
    "石川県", "福井県", "山梨県", "長野県", "岐阜県", "静岡県", "愛知県", "三重県",
    "滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県", "鳥取県", "島根県",
    "岡山県", "広島県", "山口県", "徳島県", "香川県", "愛媛県", "高知県", "福岡県",
    "佐賀県", "長崎県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
)

# 「-」等は該当数なしを表す統計表の記号であり、0 として扱う。
ZERO_SYMBOLS = frozenset({"-", "‐", "―", "－", "ー", "…", "x", "X"})

TOTAL_LABEL = "計"
OTHER_LABEL = "その他"

# 47都道府県の表記ゆれ（全角スペース、接尾辞の有無）を正規表記へ寄せる。
ALIASES: dict[str, str] = {}
for _prefecture in PREFECTURES:
    ALIASES[_prefecture] = _prefecture
    if _prefecture != "北海道":
        ALIASES[_prefecture[:-1]] = _prefecture


def normalize(value: object) -> str:
    """セル値を比較用の文字列にする。全角スペース・改行・接尾辞ゆれを吸収する。"""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", "", text)


def canonical_prefecture(value: object) -> str | None:
    return ALIASES.get(normalize(value))


def integer(value: object, *, context: str) -> int:
    if isinstance(value, str) and normalize(value) in ZERO_SYMBOLS:
        return 0
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"数値ではない値: {value!r} ({context})")
    if value < 0 or int(value) != value:
        raise ValueError(f"0以上の整数ではない値: {value!r} ({context})")
    return int(value)


def total_sheets(workbook: openpyxl.Workbook) -> list[str]:
    """「1 計」表のシート名を、出現順に返す。国立・私立・男の内訳表は対象外。"""
    pattern = re.compile(r"^1計")
    names = []
    for name in workbook.sheetnames:
        sheet = workbook[name]
        labels = (
            normalize(sheet.cell(row, column).value)
            for row in range(1, 7)
            for column in range(1, 5)
        )
        if any(pattern.match(label) for label in labels):
            names.append(name)
    if not names:
        raise ValueError(f"「1 計」表のシートが見つからない: {workbook.sheetnames}")
    return names


def locate_header_row(sheet) -> tuple[int, dict[int, str]]:
    """県名が最も多く並ぶ行をヘッダ行とし、列番号 → 出身高校所在地の対応を返す。"""
    best_row, best_origins = 0, {}
    for row in range(1, min(sheet.max_row, 10) + 1):
        origins = {}
        for column in range(1, sheet.max_column + 1):
            prefecture = canonical_prefecture(sheet.cell(row, column).value)
            if prefecture:
                origins[column] = prefecture
        if len(origins) > len(best_origins):
            best_row, best_origins = row, origins
    if len(best_origins) < 2:
        raise ValueError(f"出身高校所在地のヘッダ行を特定できない: {sheet.title}")
    return best_row, best_origins


def locate_label_column(sheet, header_row: int) -> tuple[int, dict[int, str]]:
    """県名が最も多く並ぶ列を行見出し列とし、行番号 → 大学所在地の対応を返す。"""
    best_column, best_destinations = 0, {}
    for column in range(1, min(sheet.max_column, 8) + 1):
        destinations = {}
        for row in range(header_row + 1, sheet.max_row + 1):
            prefecture = canonical_prefecture(sheet.cell(row, column).value)
            if prefecture:
                destinations[row] = prefecture
        if len(destinations) > len(best_destinations):
            best_column, best_destinations = column, destinations
    if len(best_destinations) != len(PREFECTURES):
        raise ValueError(
            f"大学所在地の行見出しが47件そろわない: {sheet.title} "
            f"({len(best_destinations)}件)"
        )
    return best_column, best_destinations


def find_labeled_column(sheet, header_row: int, label: str) -> int | None:
    for column in range(1, sheet.max_column + 1):
        if normalize(sheet.cell(header_row, column).value) == label:
            return column
    return None


def read_matrix(path: Path) -> tuple[dict[tuple[str, str], int], dict[str, int]]:
    """(出身高校所在地, 大学所在地) → 入学者数 と、大学所在地別の「その他」を返す。"""
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    matrix: dict[tuple[str, str], int] = {}
    others: dict[str, int] = {}
    row_totals: dict[str, int] = {}

    for sheet_name in total_sheets(workbook):
        sheet = workbook[sheet_name]
        header_row, origins = locate_header_row(sheet)
        _, destinations = locate_label_column(sheet, header_row)
        other_column = find_labeled_column(sheet, header_row, OTHER_LABEL)
        total_column = find_labeled_column(sheet, header_row, TOTAL_LABEL)

        for row, destination in destinations.items():
            for column, origin in origins.items():
                key = (origin, destination)
                if key in matrix:
                    raise ValueError(f"同じ組が複数シートに現れた: {key} ({path.name})")
                matrix[key] = integer(
                    sheet.cell(row, column).value,
                    context=f"{path.name}/{sheet_name} {origin} -> {destination}",
                )
            if other_column is not None:
                others[destination] = integer(
                    sheet.cell(row, other_column).value,
                    context=f"{path.name}/{sheet_name} その他 -> {destination}",
                )
            if total_column is not None:
                row_totals[destination] = integer(
                    sheet.cell(row, total_column).value,
                    context=f"{path.name}/{sheet_name} 計 -> {destination}",
                )
    workbook.close()

    missing = [
        (origin, destination)
        for origin in PREFECTURES
        for destination in PREFECTURES
        if (origin, destination) not in matrix
    ]
    if missing:
        raise ValueError(f"欠けている組がある: {len(missing)}件 例: {missing[:3]} ({path.name})")
    if len(others) != len(PREFECTURES):
        raise ValueError(f"「その他」列を読み切れていない: {len(others)}件 ({path.name})")

    # 表の「計」列と、47県 + その他 の合計が一致することを大学所在地ごとに確かめる。
    for destination in PREFECTURES:
        if destination not in row_totals:
            raise ValueError(f"「計」列が見つからない: {destination} ({path.name})")
        computed = sum(matrix[(origin, destination)] for origin in PREFECTURES)
        computed += others[destination]
        if computed != row_totals[destination]:
            raise ValueError(
                f"行合計が表の「計」と一致しない: {destination} "
                f"計算={computed} 表={row_totals[destination]} ({path.name})"
            )

    return matrix, others


def build_year(path: Path, year: int) -> tuple[list[dict[str, object]], list[dict[str, object]], int]:
    matrix, others = read_matrix(path)

    outbound: dict[str, int] = defaultdict(int)
    inbound: dict[str, int] = defaultdict(int)
    local: dict[str, int] = {}
    edges: list[dict[str, object]] = []
    for origin in PREFECTURES:
        for destination in PREFECTURES:
            movers = matrix[(origin, destination)]
            if origin == destination:
                local[origin] = movers
                continue
            outbound[origin] += movers
            inbound[destination] += movers
            if movers:
                edges.append(
                    {
                        "year": year,
                        "origin_prefecture": origin,
                        "destination_prefecture": destination,
                        "movers": movers,
                    }
                )

    nodes = []
    for prefecture in PREFECTURES:
        total_from_origin = local[prefecture] + outbound[prefecture]
        nodes.append(
            {
                "prefecture": prefecture,
                "year": year,
                "entrants_from_prefecture": total_from_origin,
                "local_entrants": local[prefecture],
                "outbound_entrants": outbound[prefecture],
                "inbound_entrants": inbound[prefecture],
                "net_inflow": inbound[prefecture] - outbound[prefecture],
                "out_of_prefecture_rate": round(outbound[prefecture] / total_from_origin, 6),
            }
        )

    # 都道府県だけに閉じたグラフでは、流出と流入の総数は一致する。
    if sum(outbound.values()) != sum(inbound.values()):
        raise AssertionError(f"都道府県間の流出・流入合計が一致しない ({path.name})")

    edges.sort(key=lambda row: int(row["movers"]), reverse=True)
    return edges, nodes, sum(others.values())


EDGE_FIELDS = ["year", "origin_prefecture", "destination_prefecture", "movers"]
NODE_FIELDS = [
    "prefecture", "year", "entrants_from_prefecture", "local_entrants",
    "outbound_entrants", "inbound_entrants", "net_inflow", "out_of_prefecture_rate",
]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def year_from_filename(path: Path) -> int:
    match = re.search(r"(\d{4})", path.stem)
    if not match:
        raise ValueError(f"ファイル名から年度を読み取れない: {path.name}。--year を指定すること")
    return int(match.group(1))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input", type=Path, nargs="*", default=None,
        help="学校基本調査の xlsx。省略時は data/raw/mext-school-basic-*.xlsx を全て処理する",
    )
    parser.add_argument("--year", type=int, help="入力が1件のときのみ、年度を明示指定できる")
    parser.add_argument("--output-dir", type=Path, default=Path("data/processed"))
    args = parser.parse_args()

    inputs = args.input or sorted(Path("data/raw").glob("mext-school-basic-*.xlsx"))
    if not inputs:
        raise SystemExit("入力ファイルがない")
    if args.year is not None and len(inputs) != 1:
        raise SystemExit("--year は入力が1件のときだけ指定できる")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    all_edges: list[dict[str, object]] = []
    all_nodes: list[dict[str, object]] = []

    for path in inputs:
        year = args.year if args.year is not None else year_from_filename(path)
        edges, nodes, other_total = build_year(path, year)
        write_csv(args.output_dir / f"od_edges_{year}.csv", EDGE_FIELDS, edges)
        write_csv(args.output_dir / f"prefectures_{year}.csv", NODE_FIELDS, nodes)
        all_edges.extend(edges)
        all_nodes.extend(nodes)
        movers = sum(int(edge["movers"]) for edge in edges)
        print(
            f"{year}: {len(edges)}辺 / 県外進学者 {movers:,}人 "
            f"/ 除外した「その他」 {other_total:,}人"
        )

    if len(inputs) > 1:
        all_edges.sort(key=lambda row: (row["year"], -int(row["movers"])))
        all_nodes.sort(key=lambda row: (row["year"], PREFECTURES.index(str(row["prefecture"]))))
        write_csv(args.output_dir / "od_edges_all.csv", EDGE_FIELDS, all_edges)
        write_csv(args.output_dir / "prefectures_all.csv", NODE_FIELDS, all_nodes)
        print(f"合算: {args.output_dir}/od_edges_all.csv, {args.output_dir}/prefectures_all.csv")


if __name__ == "__main__":
    main()
