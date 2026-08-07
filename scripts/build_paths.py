#!/usr/bin/env python3
"""学校基本調査の「都道府県別 状況別卒業者数」を、高卒進路の CSV に変換する。

出身高校所在地から見た進路の内訳を持つ。OD 表（build_od.py）と同じ主語なので、
`prefectures_*.csv` の送り出し数と同じ県・同じ年度で突き合わせられる。

年度によって表の作りが違う。

- 表は設置者別（国立・公立・私立）に分かれており、「計」のブロックは無い。
  3つを足して県の合計にする。
- 2019年度以前: 就職の内訳が「正規の職員等 / それ以外 / 一時的な仕事」。
- 2020年度以降: 「自営業主等 / 無期雇用 / 有期雇用 / 臨時労働者」に変わる。
  内訳は接続しないが、A〜D と F・G の間を全て足した「就職者等」なら通しで比較できる。
- 2020年度以前は1シートにブロックが縦に並び、2021年度以降はシートが分かれる。

そのためセル位置は決め打ちせず、ブロック見出しと列見出しの文字で位置を探す。
"""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

from build_od import PREFECTURES, canonical_prefecture, normalize, write_csv, year_from_filename

# 該当数なしを表す統計表の記号。0 として扱う。
ZERO_SYMBOLS = frozenset({"-", "‐", "―", "－", "ー", "…", "x", "X", "*"})

# 設置者のブロック見出し。「計」は男女計を指し、その後ろに設置者が続く。
FOUNDERS = ("国立", "公立", "私立")

# 列見出しの手掛かり。年度で記号が全角・半角に揺れるので正規化後の部分一致で探す。
COLUMN_KEYS = {
    "graduates": "区分",
    "university": "大学等進学者",
    "senmon": "専修学校（専門課程）進学者",
    "senshu_general": "専修学校（一般課程）等入学者",
    "training": "公共職業能力開発施設等入学者",
    "other": "左記以外の者",
    "unknown": "不詳・死亡",
}


def flat(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(value)))


def integer(value: object, *, context: str) -> int:
    if isinstance(value, str) and flat(value) in ZERO_SYMBOLS:
        return 0
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"数値ではない値: {value!r} ({context})")
    if value < 0 or int(value) != value:
        raise ValueError(f"0以上の整数ではない値: {value!r} ({context})")
    return int(value)


def find_blocks(sheet) -> list[tuple[str, int]]:
    """(設置者, 見出し行) を返す。1シートに複数ブロックが縦に並ぶ年度がある。"""
    blocks = []
    for row in range(1, sheet.max_row + 1):
        label = flat(sheet.cell(row, 1).value)
        for founder in FOUNDERS:
            # 「計国立」のように男女区分と設置者が1セルに入る。男・女のブロックは飛ばす。
            if label == f"計{founder}":
                blocks.append((founder, row))
    return blocks


def locate_columns(sheet, block_row: int) -> tuple[int, dict[str, int]]:
    """ブロック内の列見出しを読み、(データ開始行, 項目名 → 列番号) を返す。

    見出しは複数行にまたがるので、列ごとに縦へ連結してから照合する。
    """
    head_rows = range(block_row + 1, min(block_row + 6, sheet.max_row) + 1)
    stacked: dict[int, str] = {}
    for column in range(1, sheet.max_column + 1):
        stacked[column] = "".join(flat(sheet.cell(row, column).value) for row in head_rows)

    columns: dict[str, int] = {}
    for key, needle in COLUMN_KEYS.items():
        target = flat(needle)
        for column, text in stacked.items():
            if target in text:
                columns[key] = column
                break
    missing = [k for k in COLUMN_KEYS if k not in columns]
    if missing:
        raise ValueError(f"列見出しを特定できない: {missing} ({sheet.title} 行{block_row})")

    # 「区分」の列は行見出し。その右隣が「計」。
    columns["total"] = columns.pop("graduates") + 1

    first_data_row = 0
    for row in range(block_row + 1, sheet.max_row + 1):
        if canonical_prefecture(sheet.cell(row, 1).value):
            first_data_row = row
            break
    if not first_data_row:
        raise ValueError(f"都道府県の行が見つからない: {sheet.title} 行{block_row}")
    return first_data_row, columns


def read_block(sheet, block_row: int, source: str) -> dict[str, dict[str, int]]:
    """1ブロック（設置者ひとつ）の 47 県ぶんを読む。"""
    first_row, columns = locate_columns(sheet, block_row)
    # 就職は年度で内訳が変わるので、D の右から F の左までを全て足して1つにする。
    job_columns = range(columns["training"] + 1, columns["other"])

    out: dict[str, dict[str, int]] = {}
    for row in range(first_row, sheet.max_row + 1):
        prefecture = canonical_prefecture(sheet.cell(row, 1).value)
        if not prefecture:
            if len(out) >= len(PREFECTURES):
                break
            continue
        if prefecture in out:
            break
        where = f"{source} {prefecture}"
        values = {
            key: integer(sheet.cell(row, columns[key]).value, context=f"{where} {key}")
            for key in ("total", "university", "senmon", "senshu_general", "training",
                        "other", "unknown")
        }
        values["employed"] = sum(
            integer(sheet.cell(row, column).value, context=f"{where} 就職")
            for column in job_columns
        )
        out[prefecture] = values

    if len(out) != len(PREFECTURES):
        raise ValueError(f"47県そろわない: {len(out)}件 ({source} 行{block_row})")
    return out


PARTS = ("university", "senmon", "senshu_general", "training", "employed", "other", "unknown")


def build_year(path: Path, year: int) -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    totals: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    found: set[str] = set()

    for name in workbook.sheetnames:
        sheet = workbook[name]
        for founder, block_row in find_blocks(sheet):
            if founder in found:
                raise ValueError(f"設置者ブロックが重複している: {founder} ({path.name})")
            found.add(founder)
            for prefecture, values in read_block(sheet, block_row, f"{path.name}/{name}").items():
                for key, value in values.items():
                    totals[prefecture][key] += value
    workbook.close()

    if found != set(FOUNDERS):
        raise ValueError(f"設置者ブロックが欠けている: {sorted(set(FOUNDERS) - found)} ({path.name})")

    rows = []
    for prefecture in PREFECTURES:
        values = totals[prefecture]
        # 表の「計」と内訳の合計が合うことを県ごとに確かめる。
        computed = sum(values[key] for key in PARTS)
        if computed != values["total"]:
            raise ValueError(
                f"内訳の合計が「計」と一致しない: {prefecture} "
                f"計算={computed} 表={values['total']} ({path.name})"
            )
        rows.append(
            {
                "prefecture": prefecture,
                "year": year,
                "graduates": values["total"],
                "university": values["university"],
                "senmon": values["senmon"],
                "senshu_general": values["senshu_general"],
                "training": values["training"],
                "employed": values["employed"],
                "other": values["other"],
                "unknown": values["unknown"],
                "university_rate": round(values["university"] / values["total"], 6),
                "employed_rate": round(values["employed"] / values["total"], 6),
            }
        )
    return rows


FIELDS = [
    "prefecture", "year", "graduates", "university", "senmon", "senshu_general",
    "training", "employed", "other", "unknown", "university_rate", "employed_rate",
]


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input", type=Path, nargs="*", default=None,
        help="学校基本調査の xlsx。省略時は data/raw/mext-hs-paths-*.xlsx を全て処理する",
    )
    parser.add_argument("--year", type=int, help="入力が1件のときのみ、年度を明示指定できる")
    parser.add_argument("--output-dir", type=Path, default=Path("data/processed"))
    args = parser.parse_args()

    inputs = args.input or sorted(Path("data/raw").glob("mext-hs-paths-*.xlsx"))
    if not inputs:
        raise SystemExit("入力ファイルがない")
    if args.year is not None and len(inputs) != 1:
        raise SystemExit("--year は入力が1件のときだけ指定できる")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    all_rows: list[dict[str, object]] = []
    for path in inputs:
        year = args.year if args.year is not None else year_from_filename(path)
        rows = build_year(path, year)
        write_csv(args.output_dir / f"paths_{year}.csv", FIELDS, rows)
        all_rows.extend(rows)
        graduates = sum(int(r["graduates"]) for r in rows)
        university = sum(int(r["university"]) for r in rows)
        employed = sum(int(r["employed"]) for r in rows)
        print(
            f"{year}: 卒業者 {graduates:,}人 / 大学等進学 {university:,}人 "
            f"({university / graduates:.1%}) / 就職 {employed:,}人 ({employed / graduates:.1%})"
        )

    if len(inputs) > 1:
        all_rows.sort(key=lambda row: (row["year"], PREFECTURES.index(str(row["prefecture"]))))
        write_csv(args.output_dir / "paths_all.csv", FIELDS, all_rows)
        print(f"合算: {args.output_dir}/paths_all.csv")


if __name__ == "__main__":
    main()
